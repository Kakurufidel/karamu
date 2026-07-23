import logging
import csv
import uuid
from django.utils import timezone 
from django.contrib.auth import login
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.translation import gettext as _
from django.views.generic import CreateView, DetailView, ListView, UpdateView, View, FormView, DeleteView, TemplateView
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string

from .forms import EventForm, TableForm
from apps.guests.forms import RSVPForm, InvitedGuestForm
from .models import Event, EventCollaborator, Table
from apps.guests.models import InvitedGuest, GuestResponse
from apps.guests.services import TableAssignmentService, import_guests_from_excel

logger = logging.getLogger(__name__)


# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def user_can_manage_event(user, event):
    """
    Verifie si l'utilisateur peut gerer l'evenement.
    Retourne True si l'utilisateur est l'organisateur principal ou un co-organisateur accepte.
    """
    return (event.main_organizer == user or 
            EventCollaborator.objects.filter(event=event, user=user, status='accepted').exists())


# ============================================================
# LISTE DES EVENEMENTS
# ============================================================

class EventListView(LoginRequiredMixin, ListView):
    """Affiche la liste des evenements dont l'utilisateur est organisateur ou co-organisateur."""
    template_name = 'events/event_list.html'
    context_object_name = 'events'

    def get_queryset(self):
        user = self.request.user
        # Evenements dont l'utilisateur est l'organisateur principal
        owned = Event.objects.filter(main_organizer=user, is_deleted=False)
        # Evenements dont l'utilisateur est co-organisateur accepte
        coorganized = Event.objects.filter(
            collaborators__user=user,
            collaborators__status='accepted',
            is_deleted=False
        )
        return (owned | coorganized).distinct().order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_organizer'] = self.request.user.owned_events.exists()
        return context


# ============================================================
# CREATION D'EVENEMENT
# ============================================================

class EventCreateView(LoginRequiredMixin, CreateView):
    """Cree un nouvel evenement."""
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'
    success_url = reverse_lazy('authentication:dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Creer un evenement')
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        event = form.save(commit=False)
        event.main_organizer = self.request.user
        
        # Gestion des champs booleens
        event.allow_other_drinks = form.cleaned_data.get('allow_other_drinks', True)
        event.has_tables = form.cleaned_data.get('has_tables', False)
        event.is_published = form.cleaned_data.get('is_published', False)
        
        event.save()
        
        # Si la gestion des tables est activee, creer une table par defaut
        if event.has_tables:
            # ✅ CORRECTION : Supprimer 'number'
            Table.objects.create(
                event=event,
                name=_('Table principale'),
                capacity=10,
                created_by=self.request.user
            )
        
        messages.success(self.request, _('Evenement cree avec succes.'))
        return redirect(self.success_url)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)


# ============================================================
# MODIFICATION D'EVENEMENT
# ============================================================

class EventUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Modifie un evenement existant."""
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('authentication:dashboard')
    
    def test_func(self):
        event = self.get_object()
        return user_can_manage_event(self.request.user, event)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Modifier l\'evenement')
        return context

    def form_valid(self, form):
        event = form.save(commit=False)
        
        old_has_tables = self.get_object().has_tables
        new_has_tables = form.cleaned_data.get('has_tables', False)
        
        event.allow_other_drinks = form.cleaned_data.get('allow_other_drinks', True)
        event.has_tables = new_has_tables
        event.is_published = form.cleaned_data.get('is_published', False)
        
        # Si has_tables est passe de True a False, supprimer les tables
        if old_has_tables and not new_has_tables:
            event.tables.all().delete()
            messages.info(
                self.request, 
                _('Les tables ont ete supprimees car la gestion des tables a ete desactivee.')
            )
        
        # Si has_tables est passe de False a True, creer une table par defaut
        if not old_has_tables and new_has_tables:
            if not event.tables.exists():
                # ✅ CORRECTION : Supprimer 'number'
                Table.objects.create(
                    event=event,
                    name=_('Table principale'),
                    capacity=10,
                    created_by=self.request.user
                )
                messages.info(
                    self.request, 
                    _('Une table par defaut a ete creee. Vous pouvez en ajouter d\'autres.')
                )
        
        event.save()
        messages.success(self.request, _('Evenement modifie avec succes.'))
        return redirect(self.success_url)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)


# ============================================================
# DETAIL D'EVENEMENT
# ============================================================

class EventDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    Affiche les details d'un evenement.
    Inclut les statistiques, les reponses, les collaborateurs et les tables.
    """
    model = Event
    template_name = 'events/event_detail.html'
    context_object_name = 'event'
    slug_url_kwarg = 'slug'

    def test_func(self):
        event = self.get_object()
        return (event.main_organizer == self.request.user or
                event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.object

        # Recuperer les reponses
        responses = event.responses.filter(is_deleted=False)

        # Statistiques de base
        total_responses = responses.count()
        will_attend = responses.filter(will_attend=True).count()
        will_not_attend = responses.filter(will_attend=False).count()
        verified = responses.filter(verification_status='verified').count()
        unverified = responses.filter(verification_status='unverified').count()
        checkins = responses.filter(checkin_time__isnull=False).count()

        # Taux de presence
        attendance_rate = 0
        if verified > 0:
            attendance_rate = round((responses.filter(will_attend=True, verification_status='verified').count() / verified) * 100, 1)

        # Nombre de personnes attendues
        expected_guests = 0
        for r in responses.filter(will_attend=True, verification_status='verified'):
            expected_guests += r.number_of_guests

        # Nombre total d'invites pre-enregistres
        total_invited = event.invited_guests.filter(is_deleted=False).count()

        # Recuperer les tables (uniquement si has_tables est True)
        tables = event.tables.filter(is_deleted=False) if event.has_tables else []

        context.update({
            'stats': {
                'total_invited': total_invited,
                'total_responses': total_responses,
                'will_attend': will_attend,
                'will_not_attend': will_not_attend,
                'verified': verified,
                'unverified': unverified,
                'checkins': checkins,
                'attendance_rate': attendance_rate,
                'expected_guests': expected_guests,
            },
            'recent_responses': responses.order_by('-submitted_at')[:10],
            'collaborators': event.collaborators.filter(status='accepted').select_related('user'),
            'rsvp_url': event.get_rsvp_url(),
            'coorganizer_url': event.get_coorganizer_url(),
            # Variables pour les tables
            'has_tables': event.has_tables,
            'tables': tables,
            # Proprietes de l'evenement
            'display_names': event.display_names,
            'display_title': event.display_title,
            'invitation_text': event.invitation_text,
            'has_specific_names': event.has_specific_names,
            'has_tables_configured': event.has_tables_configured(),
            'drink_choices': event.get_drink_choices_with_other(),
        })
        return context


# ============================================================
# SUPPRESSION D'EVENEMENT
# ============================================================

class EventDeleteView(LoginRequiredMixin, View):
    """Supprime un evenement (soft delete)."""
    def post(self, request, *args, **kwargs):
        event = get_object_or_404(Event, slug=kwargs.get('slug'), main_organizer=request.user)
        name = event.name
        event.soft_delete()
        messages.success(request, _('L\'evenement "%(name)s" a ete supprime.') % {'name': name})
        return redirect('events:event_list')
# ============================================================
# REJOINDRE COMME CO-ORGANISATEUR
# ============================================================

class JoinCoOrganizerView(View):
    """
    Permet a un utilisateur de rejoindre un evenement comme co-organisateur.
    Si l'utilisateur est connecte, il est ajoute directement.
    Sinon, il est redirige vers la page d'inscription.
    """
    def get(self, request, slug, token):
        event = get_object_or_404(Event, slug=slug, coorganizer_token=token)
        
        if request.user.is_authenticated:
            collaborator, created = EventCollaborator.objects.get_or_create(
                event=event,
                user=request.user,
                defaults={'status': 'accepted', 'accepted_at': timezone.now()}
            )
            if not created and collaborator.status == 'pending':
                collaborator.status = 'accepted'
                collaborator.accepted_at = timezone.now()
                collaborator.save()
            
            messages.success(request, f'Vous etes maintenant co-organisateur de "{event.name}"')
            return redirect('events:event_detail', slug=event.slug)
        
        register_url = reverse('authentication:register')
        return redirect(f'{register_url}?coorganizer_code={event.coorganizer_short_code}')


class JoinCoOrganizerShortCodeView(View):
    """Permet a un utilisateur de rejoindre un evenement avec le code court."""
    def get(self, request, short_code):
        try:
            event = Event.objects.get(coorganizer_short_code=short_code.upper())
        except Event.DoesNotExist:
            messages.error(request, _('Code co-organisateur invalide.'))
            return redirect('events:event_list')

        if not request.user.is_authenticated:
            login_url = reverse('authentication:login')
            return redirect(f'{login_url}?coorganizer_code={short_code.upper()}')

        collaborator, created = EventCollaborator.objects.get_or_create(
            event=event,
            user=request.user,
            defaults={'status': 'accepted', 'accepted_at': timezone.now()}
        )
        if not created and collaborator.status == 'pending':
            collaborator.status = 'accepted'
            collaborator.accepted_at = timezone.now()
            collaborator.save()
        elif not created and collaborator.status == 'accepted':
            messages.info(request, _('Vous etes deja co-organisateur de cet evenement.'))
        else:
            messages.success(request, _('Vous etes maintenant co-organisateur de l\'evenement "%s".') % event.name)

        return redirect('events:event_detail', slug=event.slug)


# ============================================================
# FORMULAIRE RSVP (PUBLIC)
# ============================================================

class RSVPFormView(FormView):
    """Vue publique pour le formulaire RSVP."""
    template_name = 'guests/rsvp.html'
    form_class = RSVPForm

    def dispatch(self, request, *args, **kwargs):
        self.event = get_object_or_404(
            Event,
            slug=kwargs.get('slug'),
            rsvp_token=kwargs.get('token')
        )
        
        self.existing_response = None
        if request.method == 'POST':
            email = request.POST.get('email')
            if email:
                try:
                    self.existing_response = GuestResponse.objects.get(
                        event=self.event,
                        email=email,
                        is_deleted=False
                    )
                except GuestResponse.DoesNotExist:
                    pass
        
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['event'] = self.event
        
        if hasattr(self, 'existing_response') and self.existing_response:
            kwargs['instance'] = self.existing_response
        
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['drink_choices'] = self.event.get_drink_choices_with_other()
        context['allow_other_drinks'] = self.event.allow_other_drinks
        return context

    def form_valid(self, form):
        try:
            response = form.save(commit=False)
            response.event = self.event
            response.ip_address = self.request.META.get('REMOTE_ADDR')
            
            if not response.invitation_token:
                response.invitation_token = uuid.uuid4()
            
            response.save()
            response.verify_against_invited_list()
            response.send_confirmation_email()
            
            if response.will_attend:
                messages.success(self.request, 'Merci pour avoir confirme votre presence.')
            else:
                messages.info(self.request, 'Merci d\'avoir repondu.')
            
            return render(self.request, 'guests/rsvp_thanks.html', {
                'event': self.event,
                'response': response,
                'will_attend': response.will_attend,
            })
            
        except Exception as e:
            messages.error(self.request, f'Erreur lors de l\'enregistrement: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f'{field}: {error}')
        return super().form_invalid(form)


# ============================================================
# GESTION DES TABLES
# ============================================================

class TableListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Liste des tables d'un evenement avec recherche et pagination.
    Accessible uniquement si has_tables est True.
    """
    model = Table
    template_name = 'events/table_list.html'
    context_object_name = 'tables'
    paginate_by = None

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, status='accepted').exists())

    def get_paginate_by(self, queryset):
    #   per_page = self.request.GET.get('per_page')
        # if per_page and per_page.isdigit():
        #     per_page = int(per_page)
        #     if per_page in [5, 10, 20, 50]:
        #         return per_page
        return None

    def get_queryset(self):
        
        queryset = Table.objects.filter(event=self.event, is_deleted=False).prefetch_related(
            'guests',
            'invited_guests'
        )
        
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(id__icontains=search_query) |  # ✅ Utiliser id
                Q(name__icontains=search_query) |
                Q(guests__first_name__icontains=search_query) |
                Q(guests__last_name__icontains=search_query) |
                Q(invited_guests__first_name__icontains=search_query) |
                Q(invited_guests__last_name__icontains=search_query)
            ).distinct()
        
        # ✅ Utiliser 'id' pour le tri
        return queryset.order_by('id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        
        tables = self.get_queryset()
        total_guests = 0
        total_capacity = 0
        
        for table in tables:
            total_guests += table.current_guests_count
            total_capacity += table.capacity
        
        unassigned_invited = InvitedGuest.objects.filter(
            event=self.event, 
            table__isnull=True,
            is_deleted=False
        ).count()
        
        unassigned_responses = GuestResponse.objects.filter(
            event=self.event, 
            table__isnull=True, 
            will_attend=True,
            is_deleted=False
        ).count()
        
        unassigned_count = unassigned_invited + unassigned_responses
        
        context.update({
            'total_guests': total_guests,
            'total_capacity': total_capacity,
            'total_remaining': max(total_capacity - total_guests, 0),
            'unassigned_count': unassigned_count,
            'total_tables': tables.count(),
            'search_query': self.request.GET.get('q', ''),
            'current_per_page': self.get_paginate_by(self.get_queryset()),
            'has_tables': self.event.has_tables,
        })
        
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('events/table_list_ajax.html', context, request=self.request)
            return JsonResponse({
                'html': html,
                'count': context['paginator'].count if 'paginator' in context else context['tables'].count(),
                'total_guests': context['total_guests'],
                'total_capacity': context['total_capacity'],
                'total_remaining': context['total_remaining'],
                'unassigned_count': context['unassigned_count'],
                'total_tables': context['total_tables'],
            })
        return super().render_to_response(context, **response_kwargs)


class TableDetailAjaxView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vue AJAX pour recuperer les details d'une table (modal)."""

    def test_func(self):
        self.table = get_object_or_404(Table, id=self.kwargs['pk'])
        user = self.request.user
        return (self.table.event.main_organizer == user or
                self.table.event.collaborators.filter(user=user, status='accepted').exists())

    def get(self, request, pk):
        table = self.table
        guests_data = []
        
        for guest in table.guests.filter(is_deleted=False):
            status_label = 'pending'
            status_display = 'En attente'
            if guest.will_attend:
                status_label = 'confirmed'
                status_display = 'Confirme'
            elif guest.will_attend is False:
                status_label = 'declined'
                status_display = 'Refuse'
            
            guests_data.append({
                'type': 'response',
                'first_name': guest.first_name,
                'last_name': guest.last_name,
                'number_of_guests': guest.number_of_guests,
                'status': status_label,
                'status_display': status_display,
                'drink_display': guest.drink_display,
                'will_attend': guest.will_attend,
            })
        
        for invited in table.invited_guests.filter(is_deleted=False):
            has_response = GuestResponse.objects.filter(
                event=table.event,
                first_name=invited.first_name,
                last_name=invited.last_name,
                is_deleted=False
            ).exists()
            
            if not has_response:
                guests_data.append({
                    'type': 'invited',
                    'first_name': invited.first_name,
                    'last_name': invited.last_name,
                    'number_of_guests': 1,
                    'status': 'pending',
                    'status_display': 'En attente',
                    'drink_display': '-',
                    'will_attend': None,
                })
        
        status_order = {'confirmed': 0, 'pending': 1, 'declined': 2}
        guests_data.sort(key=lambda x: status_order.get(x['status'], 1))
        
        return JsonResponse({
            # ✅ Utiliser table.id comme numéro
            'number': table.id,
            'name': table.name,
            'capacity': table.capacity,
            'occupied': len(guests_data),
            'is_full': table.is_full,
            'guests': guests_data,
        })


class TableCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Cree une nouvelle table pour un evenement."""
    model = Table
    form_class = TableForm
    template_name = 'events/table_form.html'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['title'] = _('Creer une table')
        return context

    def form_valid(self, form):
        form.instance.event = self.event
        form.instance.created_by = self.request.user
        # ✅ Ne pas assigner 'number' - l'ID est généré automatiquement
        messages.success(self.request, _('Table creee avec succes.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('events:table_list', kwargs={'event_id': self.event.id})


class TableUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Modifie une table existante."""
    model = Table
    form_class = TableForm
    template_name = 'events/table_form.html'

    def test_func(self):
        table = self.get_object()
        self.event = table.event
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['title'] = _('Modifier la table')
        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Table modifiee avec succes.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('events:table_list', kwargs={'event_id': self.object.event.id})


class TableDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Supprime une table (soft delete)."""
    model = Table
    template_name = 'events/table_confirm_delete.html'

    def test_func(self):
        table = self.get_object()
        self.event = table.event
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def get_success_url(self):
        return reverse_lazy('events:table_list', kwargs={'event_id': self.object.event.id})

    def delete(self, request, *args, **kwargs):
        table = self.get_object()
        event_id = table.event.id
        table.soft_delete()
        messages.success(request, _('Table supprimee avec succes.'))
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'event_id': event_id})
        
        return redirect(self.get_success_url())


# ============================================================
# ASSIGNATION DES TABLES
# ============================================================

class AutoAssignTablesView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Assigne automatiquement les invites aux tables disponibles."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def post(self, request, event_id):
        service = TableAssignmentService(self.event)
        result = service.auto_assign_all()
        if result:
            messages.success(request, _('Les invites ont ete attribues aux tables.'))
        else:
            messages.warning(request, _('Aucune table disponible ou aucun invite a attribuer.'))
        return redirect('events:event_detail', slug=self.event.slug)


class AssignGuestTableView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Vue pour assigner ou deplacer un invite vers une table.
    Accessible uniquement si has_tables est True.
    """
    template_name = 'events/assign_guest_table.html'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        guests = GuestResponse.objects.filter(
            event=self.event,
            will_attend=True,
            is_deleted=False
        ).select_related('table').order_by('first_name', 'last_name')
        
        # ✅ Utiliser 'id' pour le tri
        tables = Table.objects.filter(
            event=self.event,
            is_deleted=False
        ).order_by('id')
        
        context = {
            'event': self.event,
            'guests': guests,
            'tables': tables,
            'selected_guest_id': request.GET.get('guest_id'),
            'has_tables': self.event.has_tables,
        }
        return render(request, self.template_name, context)

    def post(self, request, event_id):
        guest_id = request.POST.get('guest_id')
        table_id = request.POST.get('table_id')
        remove_assignment = request.POST.get('remove_assignment') == 'on'
        
        if not guest_id:
            messages.error(request, _('Veuillez selectionner un invite.'))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        try:
            guest = GuestResponse.objects.get(id=guest_id, event=self.event, is_deleted=False)
        except GuestResponse.DoesNotExist:
            messages.error(request, _('Invite introuvable.'))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        if remove_assignment:
            guest.table = None
            guest.save()
            messages.info(request, _('Assignation retiree pour {guest}.').format(guest=guest.get_full_name()))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        if not table_id:
            messages.error(request, _('Veuillez selectionner une table.'))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        try:
            table = Table.objects.get(id=table_id, event=self.event, is_deleted=False)
        except Table.DoesNotExist:
            messages.error(request, _('Table introuvable.'))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        if table.is_full:
            messages.error(request, _('Cette table est pleine (capacite: {capacity}).').format(capacity=table.capacity))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        old_table = guest.table
        guest.table = table
        guest.save()
        
        # Utiliser table.id comme numéro
        messages.success(
            request,
            _('{guest} a ete deplace(e) de la table {old} vers la table {new}.').format(
                guest=guest.get_full_name(),
                old=old_table.id if old_table else _('aucune'),
                new=table.id
            )
        )
        
        return redirect('events:assign_guest_table', event_id=event_id)

# ============================================================
# EXPORT TABLES PDF
# ============================================================

class TablesPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exporte les tables au format PDF."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            messages.warning(self.request, _('La gestion des tables n\'est pas activee pour cet evenement.'))
            return False
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from apps.guests.services import generate_tables_pdf
        pdf = generate_tables_pdf(self.event)
        if not pdf:
            messages.error(request, _('La generation du PDF a echoue.'))
            return redirect('events:event_detail', slug=self.event.slug)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="tables_{self.event.slug}.pdf"'
        return response


# ============================================================
# EXPORT TABLES CSV / EXCEL
# ============================================================

class ExportTablesCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exporte les tables au format CSV."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            return False
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        import csv
        tables = Table.objects.filter(event=self.event, is_deleted=False).prefetch_related('guests')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="tables_{self.event.slug}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'Table', 'Capacite', 'Occupee', 'Invite', 'Email', 'Telephone',
            'Boisson', 'Autre boisson', 'Verifie', 'Statut', 'Nombre de personnes'
        ])
        
        for table in tables:
            guests = table.guests.filter(will_attend=True, is_deleted=False).order_by('first_name', 'last_name')
            if guests.exists():
                for guest in guests:
                    writer.writerow([
                        table.id,  # ✅ Utiliser table.id
                        table.capacity,
                        guests.count(),
                        guest.get_full_name(),
                        guest.email or '',
                        guest.phone or '',
                        guest.drink_display,
                        guest.drink_other or '',
                        'Oui' if guest.is_verified else 'Non',
                        'Present' if guest.will_attend else 'Absent',
                        guest.number_of_guests,
                    ])
            else:
                writer.writerow([
                    table.id,
                    table.capacity,
                    0,
                    '-- Aucun invite --', '', '', '', '', '', '', ''
                ])
        
        return response


class ExportTablesExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exporte les tables au format Excel."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        if not self.event.has_tables:
            return False
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        tables = Table.objects.filter(event=self.event, is_deleted=False).prefetch_related('guests')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tables et invites"
        
        headers = [
            'Table', 'Capacite', 'Occupee', 'Invite', 'Email', 'Telephone',
            'Boisson', 'Autre boisson', 'Verifie', 'Statut', 'Nombre de personnes'
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row_idx = 2
        for table in tables:
            guests = table.guests.filter(will_attend=True, is_deleted=False).order_by('first_name', 'last_name')
            if guests.exists():
                for guest in guests:
                    ws.cell(row=row_idx, column=1, value=table.id)
                    ws.cell(row=row_idx, column=2, value=table.capacity)
                    ws.cell(row=row_idx, column=3, value=guests.count())
                    ws.cell(row=row_idx, column=4, value=guest.get_full_name())
                    ws.cell(row=row_idx, column=5, value=guest.email or '')
                    ws.cell(row=row_idx, column=6, value=guest.phone or '')
                    ws.cell(row=row_idx, column=7, value=guest.drink_display)
                    ws.cell(row=row_idx, column=8, value=guest.drink_other or '')
                    ws.cell(row=row_idx, column=9, value='Oui' if guest.is_verified else 'Non')
                    ws.cell(row=row_idx, column=10, value='Present' if guest.will_attend else 'Absent')
                    ws.cell(row=row_idx, column=11, value=guest.number_of_guests)
                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value=table.id)  # ✅ Utiliser table.id
                ws.cell(row=row_idx, column=2, value=table.capacity)
                ws.cell(row=row_idx, column=3, value=0)
                ws.cell(row=row_idx, column=4, value='-- Aucun invite --')
                row_idx += 1
        
        # Ajuster les colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tables_{self.event.slug}.xlsx"'
        wb.save(response)
        return response
    
# ============================================================
# GESTION DES CO-ORGANISATEURS
# ============================================================

class CollaboratorScanPermissionView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Active ou desactive le droit de scan d'un co-organisateur."""
    
    def test_func(self):
        collaborator = get_object_or_404(EventCollaborator, id=self.kwargs['pk'])
        return self.request.user == collaborator.event.main_organizer
    
    def post(self, request, pk):
        collaborator = get_object_or_404(EventCollaborator, id=pk)
        collaborator.can_scan = not collaborator.can_scan
        collaborator.save()
        messages.success(request, _('Droit de scan modifie.'))
        return redirect('events:event_detail', slug=collaborator.event.slug)


class CollaboratorDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Retire un co-organisateur (soft delete)."""
    model = EventCollaborator
    template_name = 'events/collaborator_confirm_delete.html'
    
    def test_func(self):
        collaborator = self.get_object()
        return self.request.user == collaborator.event.main_organizer
    
    def get_success_url(self):
        return reverse_lazy('events:event_detail', kwargs={'slug': self.object.event.slug})
    
    def delete(self, request, *args, **kwargs):
        collaborator = self.get_object()
        collaborator.soft_delete()
        messages.success(request, _('Co-organisateur retire avec succes.'))
        return redirect(self.get_success_url())