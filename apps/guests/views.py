"""
Vues de l'application guests.
Gestion des invites, reponses RSVP, check-in et exports.
"""

import csv
import json
import logging
import uuid
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils.translation import gettext as _
from django.views import View
from django.views.generic import ListView, TemplateView, FormView, CreateView
from django.utils import timezone
from django.urls import reverse_lazy
from django.conf import settings
from django.db.models import Q

from apps.events.models import Event, Table
from .models import GuestResponse, InvitedGuest
from .forms import RSVPForm, InvitedGuestForm, GuestBulkImportForm
from .services import import_guests_from_excel, generate_invitation_pdf

logger = logging.getLogger(__name__)


# ============================================================================
# 1. LISTE DES REPONSES
# ============================================================================

class GuestListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Liste des reponses des invites pour un evenement.
    """
    template_name = 'guests/guest_list.html'
    context_object_name = 'guests'

    def test_func(self):
        """Verifie que l'utilisateur peut gerer l'evenement."""
        self.event = get_object_or_404(Event, id=self.kwargs.get('event_id'))
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, status='accepted').exists())

    def get_paginate_by(self, queryset):
        """Determine le nombre d'elements par page."""
        per_page = self.request.GET.get('per_page')
        if per_page and per_page.isdigit():
            per_page = int(per_page)
            if per_page in [10, 15, 20, 30, 50, 100]:
                return per_page
        return getattr(settings, 'GUESTS_PER_PAGE', 20)

    def get_queryset(self):
        """Recupere les reponses de l'evenement."""
        return self.event.responses.all().order_by('-submitted_at')

    def get_context_data(self, **kwargs):
        """Ajoute les statistiques au contexte."""
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        responses = self.event.responses
        context['stats'] = {
            'total': responses.count(),
            'attending': responses.filter(will_attend=True).count(),
            'not_attending': responses.filter(will_attend=False).count(),
            'verified': responses.filter(verification_status='verified').count(),
            'unverified': responses.filter(verification_status='unverified').count(),
        }
        return context


# ============================================================================
# 2. LISTE DES INVITES PRE-ENREGISTRES
# ============================================================================

class InvitedGuestListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Liste des invites pre-enregistres pour un evenement.
    """
    model = InvitedGuest
    template_name = 'guests/invited_guest_list.html'
    context_object_name = 'invited_guests'
    paginate_by = 20

    def test_func(self):
        """Verifie que l'utilisateur peut gerer l'evenement."""
        self.event = get_object_or_404(Event, id=self.kwargs.get('event_id'))
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, status='accepted').exists())

    def get_queryset(self):
        """Recupere les invites avec optimisation et recherche."""
        queryset = InvitedGuest.objects.filter(
            event=self.event
        ).select_related('table')
        
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(middle_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        sort_by = self.request.GET.get('sort', 'last_name')
        if sort_by in ['first_name', 'last_name', 'email', 'created_at']:
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('last_name', 'first_name')
        
        return queryset

    def get_context_data(self, **kwargs):
        """Ajoute les statistiques et les informations de limite."""
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['current_guest_count'] = self.event.total_invited_guests()
        context['max_guests_allowed'] = self.event.max_guests_allowed
        context['remaining_guests'] = (
            self.event.max_guests_allowed - self.event.total_invited_guests()
        )
        context['can_add_guests'] = self.event.can_add_guests()[0]
        context['search_query'] = self.request.GET.get('q', '')
        context['current_sort'] = self.request.GET.get('sort', 'last_name')
        return context


# ============================================================================
# 3. AJOUT MANUEL D'UN INVITE
# ============================================================================

class AddInvitedGuestView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Ajoute manuellement un invite a la liste officielle.
    Verifie que la limite d'invites n'est pas atteinte.
    """
    model = InvitedGuest
    form_class = InvitedGuestForm
    template_name = 'guests/add_guest.html'

    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get_context_data(self, **kwargs):
        """Ajoute les informations de limite au contexte."""
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['can_add_guests'] = self.event.can_add_guests()[0]
        context['remaining_guests'] = (
            self.event.max_guests_allowed - self.event.total_invited_guests()
        )
        return context

    def form_valid(self, form):
        """
        Sauvegarde l'invite apres verification de la limite.
        """
        # Verifier la limite avant d'ajouter
        can_add, error_message = self.event.can_add_guests()
        if not can_add:
            messages.error(self.request, error_message)
            return self.form_invalid(form)
        
        form.instance.event = self.event
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Invite ajoute avec succes.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('guests:invited_list', kwargs={'event_id': self.event.id})


# ============================================================================
# 4. IMPORT EXCEL
# ============================================================================

class BulkImportGuestsView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Import Excel d'invites pre-enregistres avec verification des limites.
    """
    template_name = 'guests/bulk_import.html'

    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        """Affiche le formulaire d'import avec les informations de limite."""
        form = GuestBulkImportForm()
        current_count = self.event.total_invited_guests()
        remaining = self.event.max_guests_allowed - current_count
        
        return render(request, self.template_name, {
            'form': form,
            'event': self.event,
            'current_count': current_count,
            'max_allowed': self.event.max_guests_allowed,
            'remaining': remaining,
            'can_import': remaining > 0,
        })

    def post(self, request, event_id):
        """Traite le fichier importe avec verification des limites."""
        form = GuestBulkImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            result = import_guests_from_excel(excel_file, self.event, request.user)
            
            # Verifier si la limite a ete atteinte
            if result.get('limit_reached', False):
                error_message = (
                    result['error_messages'][0] if result['error_messages'] else ''
                )
                messages.error(request, error_message)
                return redirect('guests:invited_list', event_id=self.event.id)
            
            # Messages de succes
            if result['created'] > 0:
                messages.success(
                    request,
                    _('Import termine : %(created)s invites ajoutes, %(updated)s mis a jour.') % {
                        'created': result['created'],
                        'updated': result['updated']
                    }
                )
            elif result['updated'] > 0:
                messages.info(
                    request,
                    _('%(updated)s invites ont ete mis a jour. Aucun nouvel invite ajoute.') % {
                        'updated': result['updated']
                    }
                )
            
            # Messages d'erreur
            if result['errors'] > 0:
                messages.error(
                    request,
                    _('%(errors)s erreurs rencontrees. Details : %(details)s') % {
                        'errors': result['errors'],
                        'details': ', '.join(result['error_messages'][:3])
                    }
                )
            
            # Aucune donnee importee
            if result['created'] == 0 and result['updated'] == 0 and result['errors'] == 0:
                messages.warning(
                    request,
                    _('Aucune donnee importee. Verifiez le format du fichier.')
                )
            
            return redirect('guests:invited_list', event_id=self.event.id)
        
        messages.error(request, _('Le fichier n\'est pas valide.'))
        return render(request, self.template_name, {'form': form, 'event': self.event})


# ============================================================================
# 5. RSVP
# ============================================================================

class RSVPFormView(FormView):
    """
    Vue publique pour le formulaire RSVP.
    """
    template_name = 'guests/rsvp.html'
    form_class = RSVPForm

    def dispatch(self, request, *args, **kwargs):
        """Recupere l'evenement avant le traitement."""
        self.event = get_object_or_404(
            Event,
            slug=kwargs.get('slug'),
            rsvp_token=kwargs.get('token')
        )
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        """Passe l'evenement au formulaire."""
        kwargs = super().get_form_kwargs()
        kwargs['event'] = self.event
        return kwargs

    def get_context_data(self, **kwargs):
        """Ajoute les informations de l'evenement au contexte."""
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['drink_choices'] = self.event.get_drink_choices_with_other()
        context['allow_other_drinks'] = self.event.allow_other_drinks
        return context

    def form_valid(self, form):
        """Sauvegarde la reponse et verifie l'invite."""
        try:
            response = form.save(commit=False)
            response.event = self.event
            response.ip_address = self.request.META.get('REMOTE_ADDR')
            
            if not response.invitation_token:
                response.invitation_token = uuid.uuid4()
            
            response.save()
            response.verify_against_invited_list()
            
            if response.will_attend:
                messages.success(
                    self.request,
                    _('Merci pour avoir confirme votre presence.')
                )
            else:
                messages.info(
                    self.request,
                    _('Merci d\'avoir repondu.')
                )
            
            return render(self.request, 'guests/rsvp_thanks.html', {
                'event': self.event,
                'response': response,
                'will_attend': response.will_attend,
            })
            
        except Exception as e:
            messages.error(
                self.request,
                _('Erreur lors de l\'enregistrement: %(error)s') % {'error': str(e)}
            )
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Affiche les erreurs du formulaire."""
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f'{field}: {error}')
        return super().form_invalid(form)


class RSVPThanksView(TemplateView):
    """
    Page de remerciement apres RSVP.
    """
    template_name = 'guests/rsvp_thanks.html'


# ============================================================================
# 6. CHECK-IN
# ============================================================================

class CheckinScanView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Affiche le scanner de QR code.
    Accessible aux organisateurs et co-organisateurs avec droit de scan.
    """
    template_name = 'guests/checkin_scan.html'

    def test_func(self):
        """Verifie les droits d'acces au scanner."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, can_scan=True).exists())

    def get(self, request, event_id):
        """Affiche la page du scanner."""
        return render(request, self.template_name, {'event': self.event})


class CheckinQRView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Valide le check-in par QR code.
    """
    template_name = 'guests/checkin_qr.html'

    def test_func(self):
        """Verifie les droits d'acces au scanner."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, can_scan=True).exists())

    def post(self, request, event_id):
        """Traite le QR code scanne et valide le check-in."""
        data = request.POST.get('data')
        if not data:
            messages.error(request, _('Donnees QR code invalides.'))
            return redirect('guests:checkin_scan', event_id=event_id)
        
        try:
            guest_data = json.loads(data)
            guest = None
            
            # Recherche par short_code
            short_code = guest_data.get('short_code')
            if short_code:
                guest = GuestResponse.objects.filter(
                    short_code=short_code,
                    event=self.event
                ).first()
            
            # Recherche par nom
            if not guest:
                first_name = guest_data.get('first_name')
                last_name = guest_data.get('last_name')
                if first_name and last_name:
                    guest = GuestResponse.objects.filter(
                        event=self.event,
                        first_name=first_name,
                        last_name=last_name
                    ).first()
            
            if not guest:
                messages.error(request, _('Invite non trouve.'))
                return redirect('guests:checkin_scan', event_id=event_id)
            
            # Verifier si deja scanne
            if guest.checkin_time:
                messages.warning(
                    request,
                    _('Cet invite a deja ete scanne a {time}.').format(
                        time=guest.checkin_time.strftime('%H:%M')
                    )
                )
                return render(request, self.template_name, {
                    'guest': guest,
                    'already_checked_in': True,
                    'event': self.event,
                })
            
            # Enregistrer le check-in
            guest.checkin_time = timezone.now()
            guest.save(update_fields=['checkin_time'])
            
            messages.success(
                request,
                _('Check-in valide pour {guest} !').format(
                    guest=guest.get_full_name()
                )
            )
            
            return render(request, self.template_name, {
                'guest': guest,
                'success': True,
                'event': self.event,
            })
            
        except json.JSONDecodeError:
            messages.error(request, _('QR code invalide.'))
            return redirect('guests:checkin_scan', event_id=event_id)
        except Exception as e:
            messages.error(
                request,
                _('Erreur lors du traitement: {error}').format(error=str(e))
            )
            return redirect('guests:checkin_scan', event_id=event_id)


class CheckinManualView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Vue pour la saisie manuelle du code court.
    """
    def test_func(self):
        """Verifie les droits d'acces au scanner manuel."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, can_scan=True).exists())

    def get(self, request, event_id):
        """Traite la saisie manuelle du code court."""
        code = request.GET.get('code', '').strip().upper()
        if not code:
            messages.error(request, _('Veuillez saisir un code.'))
            return redirect('guests:checkin_scan', event_id=event_id)

        try:
            guest = GuestResponse.objects.get(short_code=code, event=self.event)
            return redirect('guests:checkin', token=guest.short_code)
        except GuestResponse.DoesNotExist:
            messages.error(request, _('Code invalide. Aucun invite trouve.'))
            return redirect('guests:checkin_scan', event_id=event_id)


class CheckInView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Vue pour scanner le QR code et valider l'arrivee d'un invite.
    """
    template_name = 'guests/checkin.html'
    success_template = 'guests/checkin_success.html'

    def test_func(self):
        """Permet l'acces a tous (la verification est faite dans dispatch)."""
        return True

    def dispatch(self, request, *args, **kwargs):
        """
        Intercepte la requete pour identifier l'invite et verifier les droits.
        """
        token = kwargs.get('token').strip('"{}')
        guest_response = None
        
        # Recherche par UUID
        try:
            uuid_obj = uuid.UUID(token)
            guest_response = GuestResponse.objects.get(invitation_token=uuid_obj)
        except (ValueError, GuestResponse.DoesNotExist):
            pass
        
        # Recherche par short_code
        if not guest_response:
            try:
                guest_response = GuestResponse.objects.get(short_code=token)
            except GuestResponse.DoesNotExist:
                pass
        
        # Recherche par ID
        if not guest_response and token.isdigit():
            try:
                guest_response = GuestResponse.objects.get(id=int(token))
            except GuestResponse.DoesNotExist:
                pass
        
        if not guest_response:
            messages.error(request, _('Code invalide. Aucun invite trouve.'))
            return redirect('guests:checkin_scan', event_id=0)
        
        self.guest_response = guest_response
        
        # Verifier les droits d'acces
        user = request.user
        event = self.guest_response.event
        is_organizer = (event.main_organizer == user)
        is_collaborator_with_scan = event.collaborators.filter(
            user=user,
            can_scan=True
        ).exists()
        
        if not (is_organizer or is_collaborator_with_scan):
            messages.error(
                request,
                _("Vous n'avez pas l'autorisation de scanner pour cet evenement.")
            )
            return redirect('events:event_list')
        
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        """Affiche les informations de l'invite pour confirmation."""
        guest = self.guest_response
        if guest.checkin_time:
            return render(request, self.template_name, {
                'guest': guest,
                'already_checked_in': True,
                'message': _("Cette invitation a deja ete scannee a {time}.").format(
                    time=guest.checkin_time.strftime('%H:%M:%S')
                )
            })
        return render(request, self.template_name, {
            'guest': guest,
            'already_checked_in': False,
        })

    def post(self, request, *args, **kwargs):
        """Valide le check-in de l'invite."""
        guest = self.guest_response
        if guest.checkin_time:
            messages.warning(request, _("Cette invitation a deja ete utilisee."))
            return redirect('guests:checkin', token=kwargs.get('token'))
        
        guest.checkin_time = timezone.now()
        guest.save(update_fields=['checkin_time'])
        
        table_number = guest.table.number if guest.table else _("non assignee")
        messages.success(
            request,
            _("Bienvenue {guest} ! Table {table}.").format(
                guest=guest.get_full_name(),
                table=table_number
            )
        )
        
        return render(request, self.success_template, {
            'guest': guest,
            'table_number': table_number,
        })


# ============================================================================
# 7. EXPORTS
# ============================================================================

class ExportGuestsCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export CSV des reponses invites.
    """
    
    def test_func(self):
        """Verifie que l'utilisateur peut gerer l'evenement."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return (self.event.main_organizer == self.request.user or
                self.event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get(self, request, event_id):
        responses = self.event.responses.all().order_by('-submitted_at')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="responses_{self.event.id}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'Prenom', 'Nom', 'Email', 'Telephone',
            'Present', 'Nombre de personnes', 'Accompagne',
            'Boisson', 'Autre boisson', 'Statut verification',
            'Soumis le'
        ])
        
        for r in responses:
            writer.writerow([
                r.first_name, r.last_name, r.email, r.phone,
                'Oui' if r.will_attend else 'Non',
                r.number_of_guests,
                'Oui' if r.is_accompanied else 'Non',
                r.drink_display, r.drink_other or '',
                r.get_verification_status_display(),
                r.submitted_at.strftime('%d/%m/%Y %H:%M'),
            ])
        return response


class ExportGuestsExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export Excel des reponses invites.
    """
    
    def test_func(self):
        """Verifie que l'utilisateur peut gerer l'evenement."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return (self.event.main_organizer == self.request.user or
                self.event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        responses = self.event.responses.all().order_by('-submitted_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reponses'
        
        headers = [
            'Prenom', 'Nom', 'Email', 'Telephone',
            'Present', 'Nombre de personnes', 'Accompagne',
            'Boisson', 'Autre boisson', 'Statut verification',
            'Soumis le'
        ]
        
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, r in enumerate(responses, 2):
            ws.cell(row=row, column=1, value=r.first_name)
            ws.cell(row=row, column=2, value=r.last_name)
            ws.cell(row=row, column=3, value=r.email)
            ws.cell(row=row, column=4, value=r.phone)
            ws.cell(row=row, column=5, value='Oui' if r.will_attend else 'Non')
            ws.cell(row=row, column=6, value=r.number_of_guests)
            ws.cell(row=row, column=7, value='Oui' if r.is_accompanied else 'Non')
            ws.cell(row=row, column=8, value=r.drink_display)
            ws.cell(row=row, column=9, value=r.drink_other or '')
            ws.cell(row=row, column=10, value=r.get_verification_status_display())
            ws.cell(row=row, column=11, value=r.submitted_at.strftime('%d/%m/%Y %H:%M'))
        
        # Ajuster les largeurs des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="responses_{self.event.id}.xlsx"'
        wb.save(response)
        return response


class ExportCheckinsCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export CSV des check-ins.
    """
    
    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        responses = GuestResponse.objects.filter(
            event=self.event,
            checkin_time__isnull=False
        ).select_related('table').order_by('checkin_time')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="checkins_{self.event.id}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Email', 'Table', "Heure d'arrivee"])
        
        for r in responses:
            writer.writerow([
                r.get_full_name(),
                r.email,
                r.table.number if r.table else '-',
                r.checkin_time.strftime('%d/%m/%Y %H:%M')
            ])
        return response


class ExportCheckinsExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export Excel des check-ins.
    """
    
    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        responses = GuestResponse.objects.filter(
            event=self.event,
            checkin_time__isnull=False
        ).select_related('table').order_by('checkin_time')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Check-ins'
        
        headers = ['Nom', 'Email', 'Table', "Heure d'arrivee"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for row, r in enumerate(responses, 2):
            ws.cell(row=row, column=1, value=r.get_full_name())
            ws.cell(row=row, column=2, value=r.email)
            ws.cell(row=row, column=3, value=r.table.number if r.table else '-')
            ws.cell(row=row, column=4, value=r.checkin_time.strftime('%d/%m/%Y %H:%M'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="checkins_{self.event.id}.xlsx"'
        wb.save(response)
        return response


class ExportInvitedCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export CSV des invites pre-enregistres.
    """

    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        guests = InvitedGuest.objects.filter(
            event=self.event
        ).order_by('last_name', 'first_name')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="invited_guests_{self.event.id}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Prenom', 'Nom', 'Postnom', 'Email', 'Telephone', 'Table'])
        for g in guests:
            writer.writerow([
                g.first_name, g.last_name, g.middle_name or '',
                g.email or '', g.phone or '',
                g.table.number if g.table else ''
            ])
        return response


class ExportInvitedExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Export Excel des invites pre-enregistres.
    """

    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        guests = InvitedGuest.objects.filter(
            event=self.event
        ).order_by('last_name', 'first_name')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Invites'
        
        headers = ['Prenom', 'Nom', 'Postnom', 'Email', 'Telephone', 'Table']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for row, g in enumerate(guests, 2):
            ws.cell(row=row, column=1, value=g.first_name)
            ws.cell(row=row, column=2, value=g.last_name)
            ws.cell(row=row, column=3, value=g.middle_name or '')
            ws.cell(row=row, column=4, value=g.email or '')
            ws.cell(row=row, column=5, value=g.phone or '')
            ws.cell(row=row, column=6, value=g.table.number if g.table else '')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="invited_guests_{self.event.id}.xlsx"'
        wb.save(response)
        return response


# ============================================================================
# 8. ASSIGNATION MANUELLE D'UNE TABLE
# ============================================================================

class AssignGuestTableView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Vue pour assigner ou deplacer un invite vers une table.
    """
    template_name = 'events/assign_guest_table.html'

    def test_func(self):
        """Verifie que l'utilisateur est l'organisateur principal."""
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        """Affiche le formulaire d'assignation des tables."""
        guests = GuestResponse.objects.filter(
            event=self.event,
            will_attend=True
        ).select_related('table').order_by('first_name', 'last_name')
        
        tables = Table.objects.filter(event=self.event).order_by('id')
        
        context = {
            'event': self.event,
            'guests': guests,
            'tables': tables,
            'selected_guest_id': request.GET.get('guest_id'),
        }
        return render(request, self.template_name, context)

    def post(self, request, event_id):
        """Traite l'assignation d'un invite a une table."""
        guest_id = request.POST.get('guest_id')
        table_id = request.POST.get('table_id')
        
        if not guest_id or not table_id:
            messages.error(
                request,
                _('Veuillez selectionner un invite et une table.')
            )
            return redirect('guests:assign_guest_table', event_id=event_id)
        
        try:
            guest = GuestResponse.objects.get(id=guest_id, event=self.event)
            table = Table.objects.get(id=table_id, event=self.event)
            
            if table.guests.count() >= table.capacity:
                messages.error(
                    request,
                    _('Cette table est pleine (capacite: {capacity}).').format(
                        capacity=table.capacity
                    )
                )
                return redirect('guests:assign_guest_table', event_id=event_id)
            
            old_table = guest.table
            guest.table = table
            guest.save()
            
            messages.success(
                request,
                _('{guest} a ete deplace(e) de la table {old} vers la table {new}.').format(
                    guest=guest.get_full_name(),
                    old=old_table.id if old_table else _('aucune'),
                    new=table.id
                )
            )
        except GuestResponse.DoesNotExist:
            messages.error(request, _('Invite introuvable.'))
        except Table.DoesNotExist:
            messages.error(request, _('Table introuvable.'))
        
        return redirect('guests:assign_guest_table', event_id=event_id)