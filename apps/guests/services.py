import csv
import io
import logging
import unicodedata
import re
from io import BytesIO
from datetime import datetime, timedelta

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils.translation import gettext as _
from django.conf import settings
from django.core.exceptions import ValidationError

from .models import InvitedGuest, GuestResponse

logger = logging.getLogger(__name__)


# ============================================================
# CONSTANTES
# ============================================================

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import qrcode
    from PIL import Image as PILImage
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed. PDF generation disabled.")


# ============================================================
# IMPORT EXCEL AVEC VÉRIFICATION DES LIMITES
# ============================================================

def import_guests_from_excel(excel_file, event, user):
    """
    Importe une liste d'invités depuis un fichier Excel avec vérification des limites.
    
    Cette fonction :
    1. Analyse le fichier Excel pour extraire les données des invités
    2. Vérifie que le nombre total d'invités ne dépasse pas la limite autorisée
    3. Importe ou met à jour les invités dans la base de données
    4. Gère automatiquement la création des tables si nécessaire
    
    Colonnes attendues dans le fichier Excel :
        - Prénom (obligatoire)
        - Nom (obligatoire) 
        - Postnom (optionnel)
        - Email (optionnel) - peut être vide
        - Téléphone (optionnel)
        - Table (optionnel) - uniquement si has_tables=True
    
    Gestion des tables :
        - Si une table avec le même nom existe déjà, elle est réutilisée
        - Sinon, une nouvelle table est créée automatiquement
        - Une table peut être attribuée à plusieurs invités
    
    Args:
        excel_file (File): Fichier Excel à importer
        event (Event): Événement auquel associer les invités
        user (User): Utilisateur effectuant l'import
    
    Returns:
        dict: Résultat de l'import avec les clés suivantes :
            - created (int): Nombre d'invités créés
            - updated (int): Nombre d'invités mis à jour
            - errors (int): Nombre d'erreurs rencontrées
            - error_messages (list): Liste des messages d'erreur détaillés
            - limit_reached (bool): True si la limite d'invités est atteinte
    
    Raises:
        ValidationError: Si la limite d'invités est dépassée
    """
    import openpyxl
    from apps.events.models import Table

    # Initialisation du résultat
    result = {
        'created': 0,
        'updated': 0,
        'errors': 0,
        'error_messages': [],
        'limit_reached': False,
    }

    try:
        # Chargement du fichier Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Lecture des en-têtes (ligne 1)
        headers = [cell.value for cell in ws[1] if cell.value]
        col_map = {}

        # Mapping des colonnes attendues
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower().strip()
                if 'prénom' in header_lower or 'prenom' in header_lower:
                    col_map['first_name'] = idx
                elif 'nom' in header_lower and 'post' not in header_lower:
                    col_map['last_name'] = idx
                elif 'postnom' in header_lower or 'post-nom' in header_lower:
                    col_map['middle_name'] = idx
                elif 'courriel' in header_lower or 'email' in header_lower:
                    col_map['email'] = idx
                elif 'téléphone' in header_lower or 'telephone' in header_lower or 'phone' in header_lower:
                    col_map['phone'] = idx
                elif 'table' in header_lower:
                    col_map['table'] = idx

        has_table_col = 'table' in col_map
        has_tables_enabled = event.has_tables

        # Vérification des colonnes obligatoires
        required_cols = ['first_name', 'last_name']
        missing_cols = [col for col in required_cols if col not in col_map]
        if missing_cols:
            result['errors'] += 1
            result['error_messages'].append(
                f"Colonnes manquantes: {', '.join(missing_cols)}. "
                "Assurez-vous d'avoir 'Prénom' et 'Nom'."
            )
            return result

        # ============================================================
        # ÉTAPE 1 : COMPTAGE DES INVITÉS À IMPORTER
        # ============================================================
        
        # Parcourir le fichier pour extraire les données et compter les invités
        guests_to_import = []
        row_errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(cell for cell in row):
                continue

            def get_cell(col_name):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ''
                return ''

            first_name = get_cell('first_name')
            last_name = get_cell('last_name')
            middle_name = get_cell('middle_name')
            email_raw = get_cell('email')
            email = email_raw if email_raw else None
            phone = get_cell('phone')
            table_name = get_cell('table') if has_table_col and has_tables_enabled else ''

            # Validation : Prénom et Nom obligatoires
            if not first_name or not last_name:
                row_errors.append(f"Ligne {row_idx}: Prénom ou nom manquant")
                continue

            # Vérifier si l'invité existe déjà (pour ne pas compter deux fois)
            exists = False
            if email:
                exists = InvitedGuest.objects.filter(event=event, email=email, is_deleted=False).exists()
            else:
                exists = InvitedGuest.objects.filter(
                    event=event,
                    first_name__iexact=first_name,
                    last_name__iexact=last_name,
                    is_deleted=False
                ).exists()
            
            guests_to_import.append({
                'row_idx': row_idx,
                'first_name': first_name,
                'last_name': last_name,
                'middle_name': middle_name,
                'email': email,
                'phone': phone,
                'table_name': table_name,
                'exists': exists,
            })

        # ============================================================
        # ÉTAPE 2 : VÉRIFICATION DE LA LIMITE D'INVITÉS
        # ============================================================
        
        # Compter les invités existants (non supprimés)
        current_guest_count = event.total_invited_guests()
        
        # Compter les nouveaux invités (ceux qui n'existent pas déjà)
        new_guests_count = sum(1 for g in guests_to_import if not g['exists'])
        
        # Vérifier si l'ajout dépasse la limite
        if current_guest_count + new_guests_count > event.max_guests_allowed:
            result['limit_reached'] = True
            error_msg = _(
                'Impossible d\'importer %(new)s invité(s). '
                'Limite actuelle : %(max)s invités. '
                'Vous avez déjà %(current)s invités. '
                'Veuillez passer à un plan supérieur pour ajouter plus d\'invités.'
            ) % {
                'new': new_guests_count,
                'max': event.max_guests_allowed,
                'current': current_guest_count,
            }
            result['errors'] += 1
            result['error_messages'].append(error_msg)
            logger.warning(
                f"Limite d'invités dépassée pour l'événement {event.id} - "
                f"Current: {current_guest_count}, Max: {event.max_guests_allowed}, "
                f"Tentative d'ajout: {new_guests_count}"
            )
            
            # Ajouter les erreurs de ligne
            result['error_messages'].extend(row_errors)
            return result

        # ============================================================
        # ÉTAPE 3 : IMPORT DES INVITÉS
        # ============================================================
        
        # Dictionnaire pour mettre en cache les tables créées pendant l'import
        tables_cache = {}
        
        for guest_data in guests_to_import:
            first_name = guest_data['first_name']
            last_name = guest_data['last_name']
            middle_name = guest_data['middle_name']
            email = guest_data['email']
            phone = guest_data['phone']
            table_name = guest_data['table_name']
            row_idx = guest_data['row_idx']

            # === GESTION DE LA TABLE ===
            table = None
            if has_tables_enabled and table_name:
                # Nettoyer le nom de la table
                table_clean = table_name.strip()
                
                # Vérifier dans le cache d'abord (optimisation)
                if table_clean in tables_cache:
                    table = tables_cache[table_clean]
                else:
                    # Chercher si la table existe déjà (par nom exact)
                    table = Table.objects.filter(
                        event=event, 
                        name__iexact=table_clean, 
                        is_deleted=False
                    ).first()
                    
                    # Si elle n'existe pas, la créer avec le nom exact
                    if not table:
                        table = Table.objects.create(
                            event=event,
                            name=table_clean,
                            capacity=10,
                            created_by=user
                        )
                        logger.info(
                            f"Table créée lors de l'import: {table_clean} "
                            f"(ID: {table.id}) pour l'événement {event.id}"
                        )
                    
                    # Ajouter au cache
                    tables_cache[table_clean] = table

            # === CRÉATION / MISE À JOUR DE L'INVITÉ ===
            try:
                if email:
                    # Recherche par email (prioritaire)
                    invited_guest, created = InvitedGuest.objects.get_or_create(
                        event=event,
                        email=email,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'middle_name': middle_name,
                            'phone': phone,
                            'created_by': user,
                            'table': table,
                        }
                    )
                    if not created:
                        # Mise à jour des informations
                        invited_guest.first_name = first_name
                        invited_guest.last_name = last_name
                        invited_guest.middle_name = middle_name
                        invited_guest.phone = phone
                        invited_guest.table = table
                        invited_guest.save()
                        result['updated'] += 1
                        logger.debug(
                            f"Invité mis à jour: {first_name} {last_name} "
                            f"(email: {email}) pour l'événement {event.id}"
                        )
                    else:
                        result['created'] += 1
                        logger.debug(
                            f"Invité créé: {first_name} {last_name} "
                            f"(email: {email}) pour l'événement {event.id}"
                        )
                else:
                    # Recherche par nom complet (sans email)
                    existing = InvitedGuest.objects.filter(
                        event=event,
                        first_name__iexact=first_name,
                        last_name__iexact=last_name,
                        email__isnull=True,
                        is_deleted=False
                    ).first()
                    
                    if existing:
                        # Mise à jour des informations (sans écraser l'email existant)
                        existing.middle_name = middle_name or existing.middle_name
                        existing.phone = phone or existing.phone
                        existing.table = table or existing.table
                        existing.save()
                        result['updated'] += 1
                        logger.debug(
                            f"Invité mis à jour: {first_name} {last_name} "
                            f"(sans email) pour l'événement {event.id}"
                        )
                    else:
                        # Création d'un nouvel invité
                        InvitedGuest.objects.create(
                            event=event,
                            first_name=first_name,
                            last_name=last_name,
                            middle_name=middle_name,
                            email=None,
                            phone=phone,
                            created_by=user,
                            table=table,
                        )
                        result['created'] += 1
                        logger.debug(
                            f"Invité créé: {first_name} {last_name} "
                            f"(sans email) pour l'événement {event.id}"
                        )
                        
            except Exception as e:
                result['errors'] += 1
                error_msg = f"Ligne {row_idx} ({first_name} {last_name}): {str(e)}"
                result['error_messages'].append(error_msg)
                logger.error(f"Erreur lors de l'import de l'invité {first_name} {last_name}: {str(e)}")

        # Ajouter les erreurs de ligne
        if row_errors:
            result['error_messages'].extend(row_errors)

    except Exception as e:
        result['errors'] += 1
        result['error_messages'].append(f"Erreur lecture fichier: {str(e)}")
        logger.error(f"Erreur lors de l'import du fichier Excel: {str(e)}", exc_info=True)

    # Logging du résultat final
    logger.info(
        f"Import terminé pour l'événement {event.id} - "
        f"{result['created']} créés, {result['updated']} mis à jour, "
        f"{result['errors']} erreurs"
    )

    return result


# ============================================================
# PDF - INVITATION
# ============================================================

class InvitationPDFService:
    """
    Service de génération d'invitation PDF avec QR code.
    
    Cette classe génère un PDF personnalisé pour chaque invité contenant :
        - Les informations de l'événement (nom, date, heure, lieu)
        - Le nom de l'invité
        - La table assignée (si disponible)
        - Un QR code pour confirmer la présence
        - Les détails du code vestimentaire (si spécifié)
    """
    
    def __init__(self, guest_response):
        """
        Initialise le service avec une réponse d'invité.
        
        Args:
            guest_response (GuestResponse): Instance de GuestResponse
        """
        self.guest_response = guest_response
        self.event = guest_response.event
    
    def generate(self):
        """
        Génère le PDF d'invitation.
        
        Returns:
            bytes: Contenu du PDF ou None en cas d'erreur
        """
        if not REPORTLAB_AVAILABLE:
            logger.error("reportlab is not installed. Cannot generate PDF.")
            return None
        
        buffer = BytesIO()
        
        try:
            # Configuration du document PDF (format A4)
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=50,
                leftMargin=50,
                topMargin=50,
                bottomMargin=50
            )
            
            styles = getSampleStyleSheet()
            story = []
            
            # ============================================================
            # STYLES PERSONNALISÉS
            # ============================================================
            
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=28,
                textColor=colors.HexColor('#2C3E50'),
                alignment=TA_CENTER,
                spaceAfter=30,
                fontName='Helvetica-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'SubtitleStyle',
                parent=styles['Heading2'],
                fontSize=18,
                textColor=colors.HexColor('#8B5CF6'),
                alignment=TA_CENTER,
                spaceAfter=20,
                fontName='Helvetica'
            )
            
            body_style = ParagraphStyle(
                'BodyStyle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#333333'),
                alignment=TA_LEFT,
                spaceAfter=8,
                fontName='Helvetica'
            )
            
            # ============================================================
            # CONTENU DU PDF
            # ============================================================
            
            story.append(Paragraph("INVITATION", title_style))
            story.append(Spacer(1, 10))
            story.append(Paragraph(self.event.name, subtitle_style))
            story.append(Spacer(1, 20))
            
            # Salutation personnalisée
            guest_name = self.guest_response.get_full_name()
            salutation = "Madame" if self.guest_response.first_name.lower() in [
                'marie', 'jeanne', 'claire', 'sophie', 'anne', 'catherine', 
                'elisabeth', 'françoise', 'nathalie', 'isabelle', 'sylvie', 
                'christine', 'laurence', 'dominique', 'valérie', 'patricia', 
                'brigitte', 'nicole', 'monique', 'micheline', 'marguerite', 
                'yolande', 'germaine', 'léontine', 'joséphine', 'mariette'
            ] else "Monsieur"
            
            story.append(Paragraph(f"{salutation} <b>{guest_name}</b>", body_style))
            story.append(Spacer(1, 10))
            
            story.append(Paragraph(
                "Nous avons le plaisir de vous inviter à notre événement :",
                body_style
            ))
            story.append(Spacer(1, 15))
            
            # Détails de l'événement
            details_data = [
                ["Date :", self.event.date.strftime('%d %B %Y') if self.event.date else 'À confirmer'],
                ["Heure :", self.event.time.strftime('%H:%M') if self.event.time else 'À confirmer'],
                ["Lieu :", self.event.location],
            ]
            if self.event.dress_code:
                details_data.append(["Tenue :", self.event.dress_code])
            
            details_table = Table(details_data, colWidths=[80, 350])
            details_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ]))
            story.append(details_table)
            story.append(Spacer(1, 20))
            
            # Table assignée (si disponible)
            if self.guest_response.table:
                table_style = ParagraphStyle(
                    'TableStyle',
                    parent=styles['Normal'],
                    fontSize=14,
                    textColor=colors.HexColor('#8B5CF6'),
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                # Utilisation de table.id comme numéro
                story.append(Paragraph(f"Table assignée : {self.guest_response.table.id}", table_style))
                story.append(Spacer(1, 15))
            
            # Message de confirmation
            story.append(Paragraph(
                "Nous sommes impatients de vous accueillir et de partager ce moment avec vous.",
                body_style
            ))
            story.append(Spacer(1, 10))
            story.append(Paragraph(
                "Veuillez confirmer votre présence en scannant le QR code ci-dessous.",
                body_style
            ))
            story.append(Spacer(1, 20))
            
            # ============================================================
            # GÉNÉRATION DU QR CODE
            # ============================================================
            
            qr = qrcode.QRCode(version=1, box_size=6, border=2)
            qr.add_data(self.guest_response.get_invitation_link())
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            qr_buffer = BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            
            qr_pil = PILImage.open(qr_buffer)
            qr_reader = ImageReader(qr_pil)
            
            qr_table = Table([[Image(qr_reader, width=60*mm, height=60*mm)]])
            qr_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(qr_table)
            story.append(Spacer(1, 10))
            
            qr_text_style = ParagraphStyle(
                'QRTextStyle',
                parent=styles['Italic'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER
            )
            story.append(Paragraph("Scannez ce QR code pour confirmer votre présence", qr_text_style))
            
            # Pied de page
            story.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#999999'),
                alignment=TA_CENTER
            )
            year = self.event.date.strftime('%Y') if self.event.date else ''
            story.append(Paragraph(
                f"© {year} {self.event.name} - Document valable uniquement avec le QR code",
                footer_style
            ))
            
            # Construction du document
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
        
        except Exception as e:
            logger.error(f"Erreur génération PDF: {str(e)}", exc_info=True)
            return None


def generate_invitation_pdf(guest_response):
    """
    Fonction wrapper pour générer un PDF d'invitation.
    
    Args:
        guest_response (GuestResponse): Réponse de l'invité
    
    Returns:
        bytes: Contenu du PDF ou None
    """
    service = InvitationPDFService(guest_response)
    return service.generate()


# ============================================================
# RAPPELS PAR EMAIL
# ============================================================

def send_reminders_for_event(event, days_before=7):
    """
    Envoie des rappels aux invités qui n'ont pas encore répondu.
    
    Args:
        event (Event): Événement concerné
        days_before (int): Nombre de jours avant l'événement
    
    Returns:
        int: Nombre de rappels envoyés
    """
    if not event.date:
        return 0
    
    today = datetime.now().date()
    event_date = event.date
    days_until = (event_date - today).days
    
    if days_until != days_before:
        return 0
    
    invited_guests = event.invited_guests.filter(is_deleted=False)
    responded_emails = event.responses.filter(
        will_attend=True,
        verification_status='verified',
        is_deleted=False
    ).values_list('email', flat=True)
    
    pending_guests = [g for g in invited_guests if g.email and g.email not in responded_emails]
    
    sent_count = 0
    for guest in pending_guests:
        try:
            response, created = GuestResponse.objects.get_or_create(
                event=event,
                email=guest.email,
                defaults={
                    'first_name': guest.first_name,
                    'last_name': guest.last_name,
                }
            )
            if response.send_reminder():
                sent_count += 1
        except Exception as e:
            logger.error(f"Erreur rappel pour {guest.email}: {str(e)}")
    
    return sent_count


# ============================================================
# TABLE ASSIGNMENT
# ============================================================

class TableAssignmentService:
    """
    Service d'assignation automatique des tables.
    
    Cette classe permet d'assigner automatiquement les invités
    aux tables disponibles en fonction de leur capacité.
    """
    
    def __init__(self, event):
        """
        Initialise le service avec un événement.
        
        Args:
            event (Event): Événement concerné
        """
        self.event = event

    def auto_assign_all(self):
        """
        Assigne automatiquement les invités aux tables disponibles.
        
        Returns:
            bool: True si au moins un invité a été assigné, False sinon
        """
        # Récupérer les tables actives (non supprimées)
        tables = list(self.event.tables.filter(is_deleted=False).order_by('id'))
        if not tables:
            logger.warning(
                f"Aucune table disponible pour l'assignation automatique "
                f"de l'événement {self.event.id}"
            )
            return False
        
        # Récupérer les invités confirmés sans table assignée
        guests = list(self.event.responses.filter(
            will_attend=True, 
            table__isnull=True,
            is_deleted=False
        ))
        if not guests:
            logger.info(
                f"Aucun invité à assigner automatiquement "
                f"pour l'événement {self.event.id}"
            )
            return False
        
        assigned_count = 0
        for guest in guests:
            for table in tables:
                if table.current_guests_count < table.capacity:
                    guest.table = table
                    guest.save()
                    assigned_count += 1
                    logger.debug(
                        f"Invité {guest.id} assigné à la table {table.id} "
                        f"pour l'événement {self.event.id}"
                    )
                    break
        
        logger.info(
            f"{assigned_count} invités assignés automatiquement "
            f"pour l'événement {self.event.id}"
        )
        return assigned_count > 0


# ============================================================
# PDF - TABLES
# ============================================================

def generate_tables_pdf(event):
    """
    Génère un PDF avec la liste des tables, invités et leurs boissons.
    
    Ce PDF est destiné aux organisateurs pour visualiser :
        - La configuration des tables
        - Les invités assignés à chaque table
        - Leurs choix de boissons (si disponibles)
        - Le statut de confirmation (confirmé, non confirmé, en attente)
    
    Args:
        event (Event): Événement concerné
    
    Returns:
        bytes: Contenu du PDF ou None en cas d'erreur
    """
    if not REPORTLAB_AVAILABLE:
        logger.error("reportlab is not installed. Cannot generate tables PDF.")
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )
    styles = getSampleStyleSheet()
    story = []
    
    # ============================================================
    # STYLES PERSONNALISÉS
    # ============================================================
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#2C3E50'),
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#7F8C8D'),
        alignment=TA_CENTER,
        spaceAfter=10,
        fontName='Helvetica'
    )
    
    table_title_style = ParagraphStyle(
        'TableTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#8B5CF6'),
        alignment=TA_LEFT,
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#95A5A6'),
        alignment=TA_LEFT,
        fontName='Helvetica'
    )
    
    # ============================================================
    # EN-TÊTE DU DOCUMENT
    # ============================================================
    
    story.append(Paragraph(f"Récapitulatif des tables - {event.name}", title_style))
    story.append(Paragraph(
        f"Date : {event.date.strftime('%d/%m/%Y')} à {event.time.strftime('%H:%M') if event.time else ''}",
        subtitle_style
    ))
    story.append(Paragraph(f"Lieu : {event.location}", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Récupérer toutes les tables (non supprimées)
    tables = event.tables.filter(is_deleted=False).order_by('id')
    
    if not tables:
        story.append(Paragraph("Aucune table configurée pour cet événement.", styles['Normal']))
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    # ============================================================
    # GÉNÉRATION DU CONTENU POUR CHAQUE TABLE
    # ============================================================
    
    for table in tables:
        # Titre de la table avec le nom exact du fichier importé
        table_display = table.name if table.name else f"Table {table.id}"
        story.append(
            Paragraph(
                f"Table {table.id} - {table_display} (capacité {table.capacity})",
                table_title_style
            )
        )
        story.append(Spacer(1, 4))
        
        # Récupération des données des invités
        guests_data = []
        
        # 1. Récupérer les GuestResponse (invités ayant répondu)
        for guest in table.guests.filter(is_deleted=False).order_by('last_name', 'first_name'):
            status = "Confirmé" if guest.will_attend else "Non confirmé"
            drink = guest.drink_display or "-"
            guests_data.append({
                'first_name': guest.first_name,
                'last_name': guest.last_name,
                'status': status,
                'drink': drink,
                'type': 'response',
                'will_attend': guest.will_attend,
            })
        
        # 2. Récupérer les InvitedGuest (invités importés n'ayant pas encore répondu)
        for invited in table.invited_guests.filter(is_deleted=False).order_by('last_name', 'first_name'):
            # Vérifier si cet invité a déjà répondu (pour éviter les doublons)
            has_response = any(
                g['first_name'].lower() == invited.first_name.lower() and 
                g['last_name'].lower() == invited.last_name.lower() 
                for g in guests_data
            )
            
            if not has_response:
                guests_data.append({
                    'first_name': invited.first_name,
                    'last_name': invited.last_name,
                    'status': 'En attente',
                    'drink': '-',
                    'type': 'invited',
                    'will_attend': None,
                })
        
        # Trier les invités : confirmés d'abord, puis en attente
        guests_data.sort(
            key=lambda x: 0 if x['will_attend'] is True 
            else (1 if x['will_attend'] is False else 2)
        )
        
        if guests_data:
            # Construction du tableau des invités
            data = [["Prénom", "Nom", "Statut", "Boisson"]]
            for g in guests_data:
                # Ajouter un indicateur visuel pour le statut
                if g['status'] == "Confirmé":
                    status_display = "✓ Confirmé"
                elif g['status'] == "Non confirmé":
                    status_display = "✗ Non confirmé"
                else:
                    status_display = "⏳ En attente"
                
                data.append([
                    g['first_name'],
                    g['last_name'],
                    status_display,
                    g['drink']
                ])
            
            # Configuration du tableau
            table_obj = Table(data, colWidths=[80, 100, 70, 100])
            table_obj.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#8B5CF6')),
                
                # Alignement
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ]))
            
            story.append(table_obj)
        else:
            story.append(Paragraph("Aucun invité assigné à cette table.", styles['Italic']))
        
        story.append(Spacer(1, 12))
    
    # ============================================================
    # PIED DE PAGE
    # ============================================================
    
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    story.append(
        Paragraph(
            f"Document généré par KaramuManage - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            footer_style
        )
    )
    
    # Construction du document
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()